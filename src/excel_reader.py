"""Read and resolve data from the Sansa monthly financial workbook."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


INCOME_SHEET = "ආදායම්"
EXPENSE_SHEET = "වියදම්"
SUMMARY_SHEET = "සාරාංශය"

LOAN_SURPLUS_LABEL = "බොල් හා අඩමාණ ණය"


@dataclass(frozen=True)
class Row:
    label: str
    value: float


@dataclass(frozen=True)
class MultiRow:
    label: str
    values: dict[date, float]
    ytd_total: float


# Income/expense detail rows that are category-summary aggregates rather than
# line items. Used to filter top-N searches.
INCOME_SUMMARY_ROWS = (2, 14, 20, 26, 31)
EXPENSE_SUMMARY_ROWS = (2, 15, 44, 47, 53)
INCOME_TOTAL_ROW = 31
EXPENSE_TOTAL_ROW = 53


def _display_text(text: object) -> str:
    """Trim whitespace but preserve Unicode joiners (ZWJ U+200D) which are
    essential for Sinhala conjunct glyphs like `න්‍ය` (yanshaya)."""
    if text is None:
        return ""
    return str(text).strip()


def _compare_key(text: object) -> str:
    """Aggressive normalization for label matching only. Strips zero-width
    joiners and whitespace so label lookups tolerate small variations."""
    if text is None:
        return ""
    return str(text).replace("\u200d", "").replace("\u200c", "").strip()


def _parse_date_cell(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _is_zero(value: object) -> bool:
    if value is None or value == "":
        return True
    try:
        return float(value) == 0.0
    except (TypeError, ValueError):
        return False


class ExcelReader:
    def __init__(self, path: Path | str):
        self.path = Path(path)
        self.wb: Workbook = load_workbook(self.path, data_only=True)

    def _sheet(self, name: str):
        return self.wb[name]

    def date_columns(self, sheet: str = INCOME_SHEET) -> list[tuple[date, str]]:
        """Return (date, column-letter) pairs for every date header in row 1.

        Skips the first column (label) and the totals column (`මුළු එකතුව`).
        Used by both detail sheets (one column per date) and summary sheet
        (paired columns — only the value column is returned).
        """
        ws = self._sheet(sheet)
        results: list[tuple[date, str]] = []
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx).value
            d = _parse_date_cell(cell)
            if d is not None:
                results.append((d, get_column_letter(col_idx)))
        return results

    def available_dates(self, sheet: str = INCOME_SHEET) -> list[date]:
        return [d for d, _ in self.date_columns(sheet)]

    def column_for(self, sheet: str, target: date) -> str:
        for d, col in self.date_columns(sheet):
            if d == target:
                return col
        raise ValueError(
            f"Date {target.isoformat()} not found in sheet {sheet!r}. "
            f"Available: {[d.isoformat() for d in self.available_dates(sheet)]}"
        )

    def latest_populated_date(self) -> date:
        """Pick the most recent date whose data column has at least one
        non-zero, non-None value in the income or expense sheet.

        Falls back to earlier months if the most recent column is empty
        (e.g. when a fresh month-end column has been added but not yet filled).
        """
        income_cols = self.date_columns(INCOME_SHEET)
        if not income_cols:
            raise ValueError(f"No date headers found in sheet {INCOME_SHEET!r}")
        income_cols.sort(key=lambda pair: pair[0], reverse=True)
        for d, col in income_cols:
            if self._has_data(INCOME_SHEET, col) or self._has_data(EXPENSE_SHEET, col):
                return d
        # Nothing populated anywhere — fall back to the latest header
        return income_cols[0][0]

    def _has_data(self, sheet: str, col_letter: str) -> bool:
        ws = self._sheet(sheet)
        col_idx = ws[col_letter + "1"].column
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if not _is_zero(value):
                return True
        return False

    def rows(
        self,
        sheet: str,
        row_range: tuple[int, int] | tuple[int, ...],
        value_col: str,
        label_col: str = "A",
    ) -> list[Row]:
        """Return non-zero rows from a range or explicit row list.

        Filters out rows whose value cell is None / "" / 0. Both top-level
        category rows and detail line items use this method.
        """
        ws = self._sheet(sheet)
        if len(row_range) == 2 and isinstance(row_range[0], int) and row_range[0] < row_range[1]:
            indices: list[int] = list(range(row_range[0], row_range[1] + 1))
        else:
            indices = list(row_range)

        out: list[Row] = []
        for r in indices:
            label = _display_text(ws[f"{label_col}{r}"].value)
            value = ws[f"{value_col}{r}"].value
            if not label or _is_zero(value):
                continue
            out.append(Row(label=label, value=float(value)))
        return out

    def loan_surplus(self, target: date) -> float:
        """Look up `බොල් හා අඩමාණ ණය` in the summary sheet by label.

        Returns the signed value at the target date's value column.
        Zero (or missing) means no slide-8 should be emitted.
        """
        ws = self._sheet(SUMMARY_SHEET)
        col = self.column_for(SUMMARY_SHEET, target)
        return self._loan_surplus_at_col(ws, col)

    def _loan_surplus_at_col(self, ws, col: str) -> float:
        target_label = _compare_key(LOAN_SURPLUS_LABEL)
        for row_idx in range(2, ws.max_row + 1):
            if _compare_key(ws.cell(row=row_idx, column=1).value) == target_label:
                value = ws[f"{col}{row_idx}"].value
                return float(value) if value is not None else 0.0
        raise ValueError(
            f"Could not find row labelled {LOAN_SURPLUS_LABEL!r} in {SUMMARY_SHEET!r}"
        )

    # --- v2 multi-month methods ----------------------------------------------

    def populated_dates(self) -> list[date]:
        """Date columns where either income or expense sheet has any non-zero
        cell. Used as the X axis for trend charts."""
        out: list[date] = []
        for d, col in self.date_columns(INCOME_SHEET):
            if self._has_data(INCOME_SHEET, col) or self._has_data(EXPENSE_SHEET, col):
                out.append(d)
        return sorted(out)

    def rows_multi(
        self,
        sheet: str,
        row_range: tuple[int, int] | tuple[int, ...],
        label_col: str = "A",
    ) -> list[MultiRow]:
        """Per-row, per-month values + YTD total — drives trend visuals."""
        ws = self._sheet(sheet)
        if len(row_range) == 2 and isinstance(row_range[0], int) and row_range[0] < row_range[1]:
            indices: list[int] = list(range(row_range[0], row_range[1] + 1))
        else:
            indices = list(row_range)

        date_cols = self.date_columns(sheet)
        ytd_col = self.ytd_column(sheet)
        out: list[MultiRow] = []
        for r in indices:
            label = _display_text(ws[f"{label_col}{r}"].value)
            if not label:
                continue
            values: dict[date, float] = {}
            for d, col in date_cols:
                v = ws[f"{col}{r}"].value
                if not _is_zero(v):
                    values[d] = float(v)
            ytd_raw = ws[f"{ytd_col}{r}"].value if ytd_col else None
            ytd_total = float(ytd_raw) if ytd_raw is not None else 0.0
            # Skip rows that are entirely zero across all months AND have no YTD
            if not values and _is_zero(ytd_total):
                continue
            out.append(MultiRow(label=label, values=values, ytd_total=ytd_total))
        return out

    def monthly_totals(self, sheet: str) -> dict[date, float]:
        """Raw total row across all date columns — does NOT apply the surplus rule."""
        ws = self._sheet(sheet)
        if sheet == INCOME_SHEET:
            row = INCOME_TOTAL_ROW
        elif sheet == EXPENSE_SHEET:
            row = EXPENSE_TOTAL_ROW
        else:
            raise ValueError(f"monthly_totals not defined for sheet {sheet!r}")
        out: dict[date, float] = {}
        for d, col in self.date_columns(sheet):
            v = ws[f"{col}{row}"].value
            if not _is_zero(v):
                out[d] = float(v)
        return out

    def loan_surplus_per_month(self) -> dict[date, float]:
        """Signed loan-surplus value per month (positive=expense, negative=income)."""
        ws = self._sheet(SUMMARY_SHEET)
        out: dict[date, float] = {}
        for d, col in self.date_columns(SUMMARY_SHEET):
            try:
                v = self._loan_surplus_at_col(ws, col)
            except ValueError:
                continue
            if v != 0:
                out[d] = v
        return out

    def adjusted_monthly_totals(self, sheet: str) -> dict[date, float]:
        """Surplus-aware totals.

        Income gains the absolute value of negative surpluses;
        expense gains positive surpluses. Reconciles to the summary
        sheet's `ශුද්ධ ලාභය` row when subtracted.
        """
        if sheet not in (INCOME_SHEET, EXPENSE_SHEET):
            raise ValueError(f"adjusted_monthly_totals: bad sheet {sheet!r}")
        raw = self.monthly_totals(sheet)
        surplus = self.loan_surplus_per_month()
        out: dict[date, float] = {}
        for d in self.populated_dates():
            base = raw.get(d, 0.0)
            s = surplus.get(d, 0.0)
            if sheet == INCOME_SHEET:
                out[d] = base + max(0.0, -s)   # negative s → income
            else:
                out[d] = base + max(0.0, s)    # positive s → expense
        return out

    def net_profit_per_month(self) -> dict[date, float]:
        income = self.adjusted_monthly_totals(INCOME_SHEET)
        expense = self.adjusted_monthly_totals(EXPENSE_SHEET)
        return {d: income.get(d, 0.0) - expense.get(d, 0.0)
                for d in self.populated_dates()}

    def ytd_column(self, sheet: str) -> str | None:
        """Return the column letter of the `මුළු එකතුව` total column, if any.

        Income/expense sheets use a single total column right after the date
        columns. Summary sheet has paired columns so we return the
        first ('value') column of the total pair.
        """
        ws = self._sheet(sheet)
        target_label = _compare_key("මුළු එකතුව")
        for col_idx in range(2, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if _compare_key(header) == target_label:
                return get_column_letter(col_idx)
        return None

    def top_line_items(
        self,
        sheet: str,
        target: date,
        n: int = 5,
        exclude_rows: tuple[int, ...] | None = None,
    ) -> list[Row]:
        """Top-N largest line items (by absolute value) for the target month.

        Excludes rows listed in `exclude_rows` (typically the
        category-summary aggregates and the grand-total row).
        """
        if exclude_rows is None:
            exclude_rows = (INCOME_SUMMARY_ROWS if sheet == INCOME_SHEET
                            else EXPENSE_SUMMARY_ROWS)
        excludes = set(exclude_rows)
        ws = self._sheet(sheet)
        col = self.column_for(sheet, target)
        items: list[Row] = []
        for r in range(2, ws.max_row + 1):
            if r in excludes:
                continue
            label = _display_text(ws[f"A{r}"].value)
            v = ws[f"{col}{r}"].value
            if not label or _is_zero(v):
                continue
            items.append(Row(label=label, value=float(v)))
        items.sort(key=lambda x: abs(x.value), reverse=True)
        return items[:n]
