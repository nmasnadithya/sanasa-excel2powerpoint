"""Generate tests/fixtures/sample.xlsx — a sanitized fixture matching the
production schema, used by the CI smoke test. Run with:

    python tests/fixtures/build_sample.py
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

DATES = [datetime(2026, 1, 31), datetime(2026, 2, 28), datetime(2026, 3, 31)]
TOTAL_HEADER = "මුළු එකතුව"


def _write_header_row(ws, label_header: str) -> None:
    ws["A1"] = label_header
    for i, d in enumerate(DATES, start=2):
        ws.cell(row=1, column=i, value=d).number_format = "yyyy-mm-dd"
    ws.cell(row=1, column=2 + len(DATES), value=TOTAL_HEADER)


def _fill_line_items(ws, rows: range, base_value: float) -> None:
    """Populate label + per-month values; column E gets the YTD sum."""
    for r in rows:
        ws.cell(row=r, column=1, value=f"පේළිය {r}")
        for col in range(2, 2 + len(DATES)):
            ws.cell(row=r, column=col, value=base_value + r + col)
        ws.cell(row=r, column=2 + len(DATES),
                value=sum(ws.cell(row=r, column=c).value
                          for c in range(2, 2 + len(DATES))))


def _aggregate_row(ws, row: int, label: str) -> None:
    ws.cell(row=row, column=1, value=label)
    # leave value cells empty — aggregate rows are excluded from top-N anyway


def build_income(ws) -> None:
    _write_header_row(ws, "විස්තරය")
    _aggregate_row(ws, 2, "කාණ්ඩ සාරාංශය 1")
    _fill_line_items(ws, range(3, 14), base_value=1000)
    _aggregate_row(ws, 14, "කාණ්ඩ සාරාංශය 2")
    _fill_line_items(ws, range(15, 20), base_value=2000)
    _aggregate_row(ws, 20, "කාණ්ඩ සාරාංශය 3")
    _fill_line_items(ws, range(21, 26), base_value=500)
    _aggregate_row(ws, 26, "කාණ්ඩ සාරාංශය 4")
    _fill_line_items(ws, range(27, 31), base_value=300)
    # Grand total at row 31 — label + per-month sum.
    ws.cell(row=31, column=1, value="මුළු ආදායම")
    for col in range(2, 2 + len(DATES)):
        ws.cell(row=31, column=col, value=sum(
            ws.cell(row=r, column=col).value or 0
            for r in list(range(3, 14)) + list(range(15, 20))
                  + list(range(21, 26)) + list(range(27, 31))
        ))
    # YTD column for total
    ws.cell(row=31, column=2 + len(DATES), value=sum(
        ws.cell(row=31, column=c).value for c in range(2, 2 + len(DATES))
    ))


def build_expense(ws) -> None:
    _write_header_row(ws, "විස්තරය")
    _aggregate_row(ws, 2, "කාණ්ඩ සාරාංශය 1")
    _fill_line_items(ws, range(3, 15), base_value=800)
    _aggregate_row(ws, 15, "කාණ්ඩ සාරාංශය 2")
    _fill_line_items(ws, range(16, 44), base_value=400)
    _aggregate_row(ws, 44, "කාණ්ඩ සාරාංශය 3")
    _fill_line_items(ws, range(45, 47), base_value=600)
    _aggregate_row(ws, 47, "කාණ්ඩ සාරාංශය 4")
    _fill_line_items(ws, range(48, 53), base_value=250)
    ws.cell(row=53, column=1, value="මුළු වියදම")
    for col in range(2, 2 + len(DATES)):
        ws.cell(row=53, column=col, value=sum(
            ws.cell(row=r, column=col).value or 0
            for r in list(range(3, 15)) + list(range(16, 44))
                  + list(range(45, 47)) + list(range(48, 53))
        ))
    ws.cell(row=53, column=2 + len(DATES), value=sum(
        ws.cell(row=53, column=c).value for c in range(2, 2 + len(DATES))
    ))


def build_summary(ws) -> None:
    _write_header_row(ws, "විස්තරය")
    # Income summary rows 4-7
    _fill_line_items(ws, range(4, 8), base_value=2500)
    # Loan surplus at row 8 with a negative value in latest month so slide-8
    # ('අධි') is exercised by the smoke test.
    ws.cell(row=8, column=1, value="බොල් හා අඩමාණ ණය")
    ws.cell(row=8, column=2, value=-100)
    ws.cell(row=8, column=3, value=-200)
    ws.cell(row=8, column=4, value=-500)
    ws.cell(row=8, column=5, value=-800)  # YTD
    # Expense summary rows 11-14
    _fill_line_items(ws, range(11, 15), base_value=1500)


def main() -> None:
    wb = Workbook()
    # openpyxl creates a default sheet — rename it to the income sheet
    income_ws = wb.active
    income_ws.title = "ආදායම්"
    build_income(income_ws)

    expense_ws = wb.create_sheet("වියදම්")
    build_expense(expense_ws)

    summary_ws = wb.create_sheet("සාරාංශය")
    build_summary(summary_ws)

    out_path = Path(__file__).parent / "sample.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
